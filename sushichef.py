#!/usr/bin/env python
import glob
import os
import sys
from ricecooker.utils import downloader, html_writer
from ricecooker.chefs import SushiChef
from ricecooker.classes import nodes, files, questions, licenses
from ricecooker.config import LOGGER              # Use LOGGER to print messages
from ricecooker.exceptions import raise_for_invalid_channel
from le_utils.constants import exercises, content_kinds, file_formats, format_presets, languages

import openpyxl


# Run constants
################################################################################
CHANNEL_ID = "d0011aa6e9e84e0f955747443f3d7e2f"             # UUID of channel
CHANNEL_NAME = "Mother Goose Club (Resources)"                           # Name of Kolibri channel
CHANNEL_SOURCE_ID = "mother-goose-club-resources"
CHANNEL_DOMAIN = "mothergooseclub.com"                         # Who is providing the content
CHANNEL_LANGUAGE = "en"                                     # Language of channel
CHANNEL_DESCRIPTION = None                                  # Description of the channel (optional)
CHANNEL_THUMBNAIL = None                                    # Local path or url to image file (optional)
CONTENT_ARCHIVE_VERSION = 1                                 # Increment this whenever you update downloaded content


# Additional constants
################################################################################
# certain topics correspond to certain filename prefixes in the resources, so we need to map those.
TOPIC_FILENAME_PREFIXES = {
    'SH Videos': ['SH.ANIM', 'SH.LIVE'],
    'Mini Books': ['Mini Book'],
    'Activity Books': ['Website.Activity Book'],
    'Board Books': ['Board Book'],
    'MGCL/MGC Anim Videos': ['MGC.ANIM', 'MGC.LIVE', 'MGC.LIVE.EPISODE'],
    'PHL Videos': ['PH.ANIM', 'PH.LIVE'],
    'MGC ABC/Counting Videos': ['MGC.ANIM', 'MGCB.2D.ANIM', 'MGCB.3D.ANIM', 'MGC.LIVE']
}

LICENSE = licenses.AllRightsLicense('Mother Goose Club')


# The chef subclass
################################################################################
class MotherGooseClubChef(SushiChef):
    """
    This class converts content from the content source into the format required by Kolibri,
    then uploads the {channel_name} channel to Kolibri Studio.
    Your command line script should call the `main` method as the entry point,
    which performs the following steps:
      - Parse command line arguments and options (run `./sushichef.py -h` for details)
      - Call the `SushiChef.run` method which in turn calls `pre_run` (optional)
        and then the ricecooker function `uploadchannel` which in turn calls this
        class' `get_channel` method to get channel info, then `construct_channel`
        to build the contentnode tree.
    For more info, see https://ricecooker.readthedocs.io
    """
    channel_info = {
        'CHANNEL_ID': CHANNEL_ID,
        'CHANNEL_SOURCE_DOMAIN': CHANNEL_DOMAIN,
        'CHANNEL_SOURCE_ID': CHANNEL_SOURCE_ID,
        'CHANNEL_TITLE': CHANNEL_NAME,
        'CHANNEL_LANGUAGE': CHANNEL_LANGUAGE,
        'CHANNEL_THUMBNAIL': CHANNEL_THUMBNAIL,
        'CHANNEL_DESCRIPTION': CHANNEL_DESCRIPTION,
    }
    DATA_DIR = os.path.abspath('chefdata')
    DOWNLOADS_DIR = os.path.join(DATA_DIR, 'downloads')
    ARCHIVE_DIR = os.path.join(DOWNLOADS_DIR, 'archive_{}'.format(CONTENT_ARCHIVE_VERSION))
    RESOURCES_DIR = os.path.join(ARCHIVE_DIR, 'Endless Learning')

    # Your chef subclass can override/extend the following method:
    # get_channel: to create ChannelNode manually instead of using channel_info
    # pre_run: to perform preliminary tasks, e.g., crawling and scraping website
    # __init__: if need to customize functionality or add command line arguments
    def construct_channel(self, *args, **kwargs):
        """
        Creates ChannelNode and build topic tree
        Args:
          - args: arguments passed in on the command line
          - kwargs: extra options passed in as key="value" pairs on the command line
            For example, add the command line option   lang="fr"  and the value
            "fr" will be passed along to `construct_channel` as kwargs['lang'].
        Returns: ChannelNode
        """
        channel = self.get_channel(*args, **kwargs)  # Create ChannelNode from data in self.channel_info

        self.load_content_from_spreadsheet()

        for type in self.content_by_type:
            topic = nodes.TopicNode(source_id=type, title=type)

            channel.add_child(topic)
            for content_node in self.content_by_type[type]:
                full_path = os.path.join(self.RESOURCES_DIR, content_node['file'])
                if os.path.splitext(full_path)[1] == '.mp4':
                    node = nodes.VideoNode(source_id=content_node['file'], title=content_node['title'], license=LICENSE)
                    node.add_file(files.VideoFile(full_path))
                    topic.add_child(node)
                else:
                    node = nodes.DocumentNode(source_id=content_node['file'], title=content_node['title'], license=LICENSE)
                    node.add_file(files.DocumentFile(full_path))
                    topic.add_child(node)

        return channel

    def load_content_from_spreadsheet(self):
        wb = openpyxl.load_workbook(os.path.join(self.ARCHIVE_DIR, "Resources.xlsx"))
        sheet = wb.active

        num_rows = sheet.max_row
        num_cols = sheet.max_column

        self.col_names = []
        self.content_by_type = {}

        for c in range(1, num_cols + 1): 
            cell_obj = sheet.cell(row = 1, column = c)
            self.col_names.append(cell_obj.value or '')

        for r in range(2, num_rows + 1):
            for c in range(1, num_cols + 1):
                if not self.col_names[c-1]:
                    continue

                type = self.col_names[c-1].strip()
                cell_obj = sheet.cell(row = r, column = c)
                value = cell_obj.value
                resources = os.listdir(self.RESOURCES_DIR)
                if value:
                    value = value.strip()
                    title = value
                    resource_file = None
                    prefixes = []
                    prefix_list = TOPIC_FILENAME_PREFIXES[type]
                    if '(2D Anim)' in value:
                        prefix_list = ['MGCB.2D.ANIM']
                    elif '(3D Anim)' in value:
                        prefix_list = ['MGCB.3D.ANIM']

                    for prefix in prefix_list:
                        anim_str = '(Anim)'
                        live_str = '(Live)'
                        if anim_str in value and "LIVE" in prefix:
                            continue
                        
                        if live_str in value and "ANIM" in prefix:
                            continue

                        for remove in [anim_str, live_str, '(2D Anim)', '(3D Anim)']:
                            value = value.replace(remove, '').strip()
                        
                        file_prefix = "{}.{}.".format(prefix, value)
                        prefixes.append(file_prefix)
                        found_files = []
                        for resource in resources:
                            # make sure we do case insensitive search.
                            if resource.lower().startswith(file_prefix.lower()):
                                found_files.append(resource)
                            else:
                                replacements = [[',', ''], [' Group', '.Group'], [' Noa', '.Noa'], [' Robert', '.Robert'], [' Caralyn', '.Caralyn']]
                                for text, replacement in replacements:
                                    if resource.lower().startswith(file_prefix.replace(text, replacement).lower()):
                                        resource_file = resource
                        for found_file in found_files:
                            # HACK Alert! Because filenames aren't an exact match to what is in the spreadsheet, we guess based on prefix, which
                            # may return multiple results. If we get multiple results, we want the one that most closely matched the prefix,
                            # in other words, the smallest number of characters that matched.
                            if not resource_file or len(found_file) < len(resource_file):
                                # There's one MOV file in the resources for some reason, and it also has an MP4, so just ignore it.
                                if found_file.lower().endswith('.mov'):
                                    continue
                                resource_file = found_file


                    if resource_file:
                        if not type in self.content_by_type:
                            self.content_by_type[type] = []
                        content = {'title': title, 'file': resource_file}
                        print("Adding content {}".format(content))
                        self.content_by_type[type].append(content)
                    else:
                        print("Unable to find file for {}".format(prefixes))



# CLI
################################################################################
if __name__ == '__main__':
    # This code runs when sushichef.py is called from the command line
    chef = MotherGooseClubChef()
    chef.main()